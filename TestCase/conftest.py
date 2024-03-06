"""
#!/usr/bin/env python
# -*- coding:utf-8 -*-
@File : conftest.py
@Author : putongrenji
@Time : 2022/5/4 11:17
@Motto:Don't ever let somebody tell  you can't do something
"""
import coloredlogs
import openpyxl
import pytest
from utils.logger_util import LoggerUtil


@pytest.fixture(name='api_log')
def api_log():
    # print(request.param)
    lu = LoggerUtil()
    logger = lu.get_logger()
    logger.handlers.clear()
    coloredlogs.install(level='DEBUG', logger=logger,
                        level_styles={'info': {'color': 'green'},
                                      'warning': {'color': 'yellow'},
                                      'error': {'color': 'red'},
                                      'critical': {'color': 'red', 'bold': True}})
    logger.info("----------测试用例执行开始----------")
    yield logger
    # out = yield logger
    # report = out.get_result()
    # print(('运行结果: %s' % report.outcome))
    logger.info("----------测试用例执行结束----------\n")
    lu.remove_handler()


def write_case_result(excel_path, sheet_name, report):
    # 打开工作簿
    wb = openpyxl.load_workbook(excel_path)
    # 获取sheet对象
    sheet = wb[sheet_name]

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        # 使用索引来访问行的数据
        if row[0].value == report.nodeid:
            row[8].value = report.outcome

    # 保存工作簿
    wb.save(excel_path)
    wb.close()


if __name__ == '__main__':
    pass