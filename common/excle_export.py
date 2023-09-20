import openpyxl


def excle_read(excle_url, sheet_name):
    # 加载工作薄
    wb = openpyxl.load_workbook(excle_url)
    # 获取sheet对象
    sheet = wb[sheet_name]
    # print(sheet)
    # 获取最大行和最大列
    all_list = []
    # all_list.append(sheet_name)
    for row in range(2, sheet.max_row+1):
        row_list = []
        for column in range(1, sheet.max_column+1):
            row_list.append(sheet.cell(row, column).value)
        all_list.append(row_list)
    # print(all_list)
    # print(sheet_name)
    return all_list


def read_all_excel_data(excel_url):
    data = {}
    # 加载工作簿
    workbook = openpyxl.load_workbook(excel_url, data_only=True)

    for sheet_name in workbook.sheetnames:
        sheet_data = excle_read(excel_url, sheet_name)
        data[sheet_name] = sheet_data
    return data


# 重写状态
def write_case_result(excel_path, sheet_name, case_num, case_status):
    # 打开工作簿
    wb = openpyxl.load_workbook(excel_path)
    # 获取sheet对象
    sheet = wb[sheet_name]

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        # 使用索引来访问行的数据
        # 初始化状态为 'failed'
        status = case_status
        if row[0].value == case_num:
            row[8].value = status


    # 保存工作簿
    wb.save(excel_path)
    wb.close()


if __name__ == '__main__':
    # print(excle_read("../data/ele.xlsx", "Sheet1"))
    # print(read_all_excel_data('../data/api_test.xlsx'))
    print(excle_read('../data/api_test.xlsx', 'Sheet1'))

